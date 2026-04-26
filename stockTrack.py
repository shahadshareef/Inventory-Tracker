from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
from tkinter import filedialog

class StockManagerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Inventory Manager")
        self.root.geometry("500x550")

        self.inventory = {}
        self.file = "inventory.json"

        self.load_inventory()

        # --- Style ---
        style = ttk.Style()
        style.theme_use("clam")

        # --- Layout Config ---
        root.columnconfigure(0, weight=1)
        root.rowconfigure(3, weight=1)

        # --- Top Frame ---
        top_frame = ttk.Frame(root, padding=10)
        top_frame.grid(row=0, column=0, sticky="ew")

        ttk.Label(top_frame, text="Item Name").grid(row=0, column=0, sticky="w")
        self.item_entry = ttk.Entry(top_frame)
        self.item_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        ttk.Label(top_frame, text="Quantity (+/-)").grid(row=0, column=1, sticky="w")
        self.qty_entry = ttk.Entry(top_frame)
        self.qty_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        top_frame.columnconfigure(0, weight=1)
        top_frame.columnconfigure(1, weight=1)

        # --- Buttons ---
        btn_frame = ttk.Frame(root, padding=10)
        btn_frame.grid(row=1, column=0, sticky="ew")

        ttk.Button(btn_frame, text="Update Stock", command=self.update_stock).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="Clear", command=self.clear_inputs).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Delete Item", command=self.delete_item).grid(row=0, column=2, padx=5)
        ttk.Button(btn_frame, text="Import Excel", command=self.import_excel).grid(row=0, column=3, padx=5)

        # --- Search ---
        search_frame = ttk.Frame(root, padding=10)
        search_frame.grid(row=2, column=0, sticky="ew")

        ttk.Label(search_frame, text="Search").pack(anchor="w")
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(fill="x", pady=5)

        # Suggestion dropdown
        self.suggestion_box = tk.Listbox(root, height=5)
        self.suggestion_box.bind("<<ListboxSelect>>", self.select_suggestion)

        self.search_entry.bind("<KeyRelease>", self.on_search)

        # --- Filter ---
        self.filter_var = tk.StringVar(value="All")

        filter_frame = ttk.Frame(root, padding=10)
        filter_frame.grid(row=3, column=0, sticky="ew")

        ttk.Label(filter_frame, text="Filter:").pack(side="left")

        ttk.Combobox(
            filter_frame,
            textvariable=self.filter_var,
            values=["All", "Low Stock (<5)", "In Stock"],
            state="readonly"
        ).pack(side="left", padx=5)

        self.filter_var.trace_add("write", lambda *args: self.update_inventory_display())

        # --- Inventory List ---
        list_frame = ttk.Frame(root, padding=10)
        list_frame.grid(row=4, column=0, sticky="nsew")

        root.rowconfigure(4, weight=1)

        self.inventory_list = tk.Listbox(list_frame)
        self.inventory_list.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.inventory_list.yview)
        scrollbar.pack(side="right", fill="y")

        self.inventory_list.config(yscrollcommand=scrollbar.set)
        self.inventory_list.bind("<<ListboxSelect>>", self.on_select)

        # --- Shortcuts ---
        root.bind("<Return>", lambda e: self.update_stock())
        root.bind("<Delete>", lambda e: self.delete_item())

        self.update_inventory_display()

    # core logic

    def import_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, qty = row

                if not name or not isinstance(qty, int):
                    continue

                current = self.inventory.get(name, 0)
                new_qty = current + qty

                if new_qty < 0:
                    new_qty = 0

                if new_qty == 0:
                    self.inventory.pop(name, None)
                else:
                    self.inventory[name] = new_qty

            self.save_inventory()
            self.update_inventory_display()

            messagebox.showinfo("Success", "Excel data imported successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to import file:\n{e}")

    def update_stock(self):
        name = self.item_entry.get().strip().title()
        qty_text = self.qty_entry.get().strip()

        # Validate input
        if not name or not qty_text:
            messagebox.showwarning(
                "Input Error",
                "Enter item name and quantity."
            )
            return

        try:
            qty = int(qty_text)
        except ValueError:
            messagebox.showwarning(
                "Input Error",
                "Quantity must be a valid number."
            )
            return

        # Update stock
        new_qty = self.inventory.get(name, 0) + qty

        # Prevent negative stock
        if new_qty < 0:
            messagebox.showerror(
                "Error",
                "Stock cannot go below 0."
            )
            return

        # Remove item if stock becomes 0
        if new_qty == 0:
            self.inventory.pop(name, None)
        else:
            self.inventory[name] = new_qty

        # Save + refresh UI
        self.save_inventory()
        self.update_inventory_display()
        self.clear_inputs()

        messagebox.showinfo(
        "Success",
        f"{name} stock updated successfully."
    )

    def delete_item(self):
        name = self.item_entry.get().strip()

        if name in self.inventory:
            if messagebox.askyesno("Confirm", f"Delete {name}?"):
                del self.inventory[name]
                self.save_inventory()
                self.update_inventory_display()
                self.clear_inputs()
        else:
            messagebox.showerror("Error", "Item not found.")

    # Search 

    def on_search(self, event=None):
        self.show_suggestions()
        self.update_inventory_display()

    def show_suggestions(self):
        text = self.search_entry.get().lower()

        if not text:
            self.suggestion_box.place_forget()
            return

        matches = [n for n in self.inventory if text in n.lower()]

        if not matches:
            self.suggestion_box.place_forget()
            return

        self.suggestion_box.delete(0, tk.END)

        for item in matches[:5]:
            self.suggestion_box.insert(tk.END, item)

        x = self.search_entry.winfo_x()
        y = self.search_entry.winfo_y() + self.search_entry.winfo_height()

        self.suggestion_box.place(in_=self.search_entry.master, x=x, y=y, width=self.search_entry.winfo_width())

    def select_suggestion(self, event):
        selection = self.suggestion_box.curselection()
        if selection:
            value = self.suggestion_box.get(selection[0])
            self.search_entry.delete(0, tk.END)
            self.search_entry.insert(0, value)
            self.suggestion_box.place_forget()
            self.update_inventory_display()

    

    def update_inventory_display(self, event=None):
        search = self.search_entry.get().lower()
        filter_mode = self.filter_var.get()

        self.inventory_list.delete(0, tk.END)

        for name, qty in sorted(self.inventory.items()):

            if search and search not in name.lower():
                continue

            if filter_mode == "Low Stock (<5)" and qty >= 5:
                continue
            elif filter_mode == "In Stock" and qty <= 0:
                continue

            index = self.inventory_list.size()
            self.inventory_list.insert(tk.END, f"{name}: {qty}")

            if qty < 5:
                self.inventory_list.itemconfig(index, {'fg': 'red'})

    def on_select(self, event):
        selection = self.inventory_list.curselection()
        if selection:
            selected = self.inventory_list.get(selection[0])
            if ":" in selected:
                name = selected.split(":")[0]
                self.item_entry.delete(0, tk.END)
                self.item_entry.insert(0, name)

    def clear_inputs(self):
        self.item_entry.delete(0, tk.END)
        self.qty_entry.delete(0, tk.END)

    

    def save_inventory(self):
        with open(self.file, "w") as f:
            json.dump(self.inventory, f)

    def load_inventory(self):
        if os.path.exists(self.file):
            with open(self.file, "r") as f:
                self.inventory = json.load(f)


if __name__ == "__main__":
    root = tk.Tk()
    app = StockManagerGUI(root)
    root.mainloop()